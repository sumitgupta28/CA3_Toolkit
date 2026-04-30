"""
STEP 3 — Extract student data from all .docx files and generate marks.xlsx
---------------------------------------------------------------------------
Usage:
    python extract_to_excel.py

Expects all student .docx files to be in the 'students/' folder.
Produces 'marks.xlsx' with one row per student.
You then fill in the Marks Awarded + Feedback columns and run fill_marks_and_export.py.
"""

import os
import sys
import glob
from pathlib import Path

# ── dependency check ────────────────────────────────────────────────────────
try:
    from docx import Document
except ImportError:
    print("ERROR: python-docx not found. Run:  pip install python-docx openpyxl")
    sys.exit(1)
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not found. Run:  pip install python-docx openpyxl")
    sys.exit(1)

# ── configuration ────────────────────────────────────────────────────────────
STUDENTS_FOLDER = "students"          # folder containing student .docx files
OUTPUT_FOLDER   = "CA3_Marks"         # output folder
OUTPUT_EXCEL    = "marks.xlsx"        # output spreadsheet

# Top-table field labels → Excel column names
# Left-cell labels (row 0-5, col 0) and right-cell labels (row 0-5, col 1)
TOP_TABLE_MAP = {
    # (row_index, col_index): excel_column_name
    (0, 0): "program_name",
    (0, 1): "year_semester",
    (1, 0): "subject",
    (1, 1): "paper_code",
    (2, 0): "upid",
    (2, 1): "date_of_examination",
    (3, 0): "student_name",
    (3, 1): "roll_number",
    (4, 0): "subject_teacher",
    (4, 1): "mobile_number",
    (5, 0): "full_marks",
    (5, 1): "duration",
}

# Question rows in the Marks Tabulation table (row index 1 onward = data rows)
QUESTION_LABELS = ["1a", "1b", "1c", "1d", "1e", "2", "3", "4", "5", "6", "7"]


def get_cell_text(cell):
    """Return clean text from a table cell, stripping label prefix."""
    return cell.text.strip()


def extract_value_after_colon(text):
    """'Program Name: BCA' → 'BCA'  |  'Program Name:' → ''"""
    if ":" in text:
        return text.split(":", 1)[1].strip()
    return text.strip()


def parse_student_doc(filepath):
    """
    Open a student .docx and extract:
      - all fields from the top info table
      - marks_allotted for each question (col 1 of marks table)
    Returns a dict, or None on error.
    """
    try:
        doc = Document(filepath)
    except Exception as e:
        msg = str(e)
        if "content type" in msg or "not a Word file" in msg:
            print(f"  ⚠  Corrupt/invalid file skipped: {Path(filepath).name}")
            print(f"       The file is not a valid Word document ({msg.split('content type')[-1].strip() if 'content type' in msg else msg})")
            print(f"       Ask the student to re-export from Microsoft Word and resubmit.")
        else:
            print(f"  ⚠  Could not open {Path(filepath).name}: {e}")
        return None

    data = {"source_file": Path(filepath).name}

    tables = doc.tables
    if len(tables) < 1:
        print(f"  ⚠  No tables found in {filepath}")
        return data

    # ── Table 0: top info table (6 rows × 2 cols) ────────────────────────
    info_table = tables[0]
    for (row_i, col_i), col_name in TOP_TABLE_MAP.items():
        try:
            raw = info_table.rows[row_i].cells[col_i].text.strip()
            data[col_name] = extract_value_after_colon(raw)
        except IndexError:
            data[col_name] = ""

    # ── Table 2: marks tabulation (header row + 11 data rows) ────────────
    # Table 0 = info, Table 1 = Assessment Rubrics (untouched), Table 2 = marks
    if len(tables) >= 3:
        marks_table = tables[2]
        for i, label in enumerate(QUESTION_LABELS):
            row_i = i + 1          # row 0 is the header
            try:
                row = marks_table.rows[row_i]
                cells = row.cells
                data[f"marks_allotted_{label}"]  = cells[1].text.strip() if len(cells) > 1 else ""
                data[f"marks_awarded_{label}"]   = cells[2].text.strip() if len(cells) > 2 else ""
                data[f"course_outcome_{label}"]  = cells[3].text.strip() if len(cells) > 3 else ""
                data[f"blooms_level_{label}"]    = cells[4].text.strip() if len(cells) > 4 else ""
                data[f"remarks_{label}"]         = cells[5].text.strip() if len(cells) > 5 else ""
                data[f"ar_reference_{label}"]    = cells[6].text.strip() if len(cells) > 6 else ""
            except IndexError:
                for key in ["marks_allotted", "marks_awarded", "course_outcome",
                            "blooms_level", "remarks", "ar_reference"]:
                    data[f"{key}_{label}"] = ""
    else:
        for label in QUESTION_LABELS:
            for key in ["marks_allotted", "marks_awarded", "course_outcome",
                        "blooms_level", "remarks", "ar_reference"]:
                data[f"{key}_{label}"] = ""

    return data


def build_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Entry"

    # ── column order ─────────────────────────────────────────────────────
    info_cols = [
        "source_file", "upid", "student_name", "roll_number",
        "program_name", "year_semester", "subject", "paper_code",
        "date_of_examination", "subject_teacher", "mobile_number",
        "full_marks", "duration",
    ]
    marks_allotted_cols  = [f"marks_allotted_{q}"  for q in QUESTION_LABELS]
    course_outcome_cols  = [f"course_outcome_{q}"  for q in QUESTION_LABELS]
    blooms_level_cols    = [f"blooms_level_{q}"    for q in QUESTION_LABELS]
    marks_awarded_cols   = [f"marks_awarded_{q}"   for q in QUESTION_LABELS]
    remarks_cols         = [f"remarks_{q}"         for q in QUESTION_LABELS]
    ar_reference_cols    = [f"ar_reference_{q}"    for q in QUESTION_LABELS]
    feedback_cols = ["strengths", "areas_for_improvement", "corrective_measures"]

    all_cols = (info_cols + marks_allotted_cols + course_outcome_cols
                + blooms_level_cols + marks_awarded_cols
                + remarks_cols + ar_reference_cols + feedback_cols)

    # ── human-readable headers ────────────────────────────────────────────
    header_labels = {
        "source_file": "Source File",
        "upid": "UPID",
        "student_name": "Student Name",
        "roll_number": "Roll Number",
        "program_name": "Program Name",
        "year_semester": "Year/Semester",
        "subject": "Subject",
        "paper_code": "Paper Code",
        "date_of_examination": "Date of Exam",
        "subject_teacher": "Subject Teacher",
        "mobile_number": "Mobile",
        "full_marks": "Full Marks",
        "duration": "Duration",
    }
    for q in QUESTION_LABELS:
        header_labels[f"marks_allotted_{q}"]  = f"Allotted {q.upper()}"
        header_labels[f"course_outcome_{q}"]  = f"Course Outcome {q.upper()}"
        header_labels[f"blooms_level_{q}"]    = f"Blooms Level {q.upper()}"
        header_labels[f"marks_awarded_{q}"]   = f"Awarded {q.upper()}"
        header_labels[f"remarks_{q}"]         = f"Remarks {q.upper()}"
        header_labels[f"ar_reference_{q}"]    = f"AR Ref {q.upper()}"
    header_labels["strengths"]              = "Strengths"
    header_labels["areas_for_improvement"]  = "Areas for Improvement"
    header_labels["corrective_measures"]    = "Corrective Measures"

    # ── styles ────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill_info     = PatternFill("solid", fgColor="1F3864")   # dark blue
    header_fill_allotted = PatternFill("solid", fgColor="375623")   # dark green
    header_fill_outcome  = PatternFill("solid", fgColor="1F6B6B")   # dark teal  (Course Outcome)
    header_fill_blooms   = PatternFill("solid", fgColor="17537A")   # dark cyan  (Bloom's Level)
    header_fill_awarded  = PatternFill("solid", fgColor="833C00")   # dark orange
    header_fill_remarks  = PatternFill("solid", fgColor="6B1F1F")   # dark maroon (Remarks / AR Ref)
    header_fill_feedback = PatternFill("solid", fgColor="4B0082")   # dark purple

    row_fill_even = PatternFill("solid", fgColor="F2F2F2")

    def header_fill_for(col):
        if col in info_cols:           return header_fill_info
        if col in marks_allotted_cols: return header_fill_allotted
        if col in course_outcome_cols: return header_fill_outcome
        if col in blooms_level_cols:   return header_fill_blooms
        if col in marks_awarded_cols:  return header_fill_awarded
        if col in remarks_cols:        return header_fill_remarks
        if col in ar_reference_cols:   return header_fill_remarks
        return header_fill_feedback

    # ── write header row ──────────────────────────────────────────────────
    for c_idx, col in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header_labels.get(col, col))
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill_for(col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # ── write data rows ───────────────────────────────────────────────────
    for r_idx, row_data in enumerate(rows, start=2):
        fill = row_fill_even if r_idx % 2 == 0 else None
        for c_idx, col in enumerate(all_cols, start=1):
            val  = row_data.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = border
            if fill:
                cell.fill = fill
            # yellow = examiner fills; light teal = pre-filled from template
            if col in marks_awarded_cols or col in remarks_cols or col in ar_reference_cols or col in feedback_cols:
                cell.fill = PatternFill("solid", fgColor="FFFACD")
            elif col in course_outcome_cols or col in blooms_level_cols:
                cell.fill = PatternFill("solid", fgColor="E0F4F4")

    # ── column widths ─────────────────────────────────────────────────────
    col_widths = {
        "source_file": 28, "upid": 14, "student_name": 22,
        "roll_number": 14, "program_name": 20, "year_semester": 14,
        "subject": 22, "paper_code": 16, "date_of_examination": 16,
        "subject_teacher": 20, "mobile_number": 14,
        "full_marks": 12, "duration": 12,
        "strengths": 30, "areas_for_improvement": 30, "corrective_measures": 30,
    }
    for q in QUESTION_LABELS:
        col_widths[f"course_outcome_{q}"] = 14
        col_widths[f"blooms_level_{q}"]   = 14
        col_widths[f"remarks_{q}"]        = 18
        col_widths[f"ar_reference_{q}"]   = 14
    for col in all_cols:
        col_letter = openpyxl.utils.get_column_letter(all_cols.index(col) + 1)
        ws.column_dimensions[col_letter].width = col_widths.get(col, 13)

    # ── freeze top row ────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    folder = Path(STUDENTS_FOLDER)
    if not folder.exists():
        print(f"ERROR: Folder '{STUDENTS_FOLDER}' not found.")
        print("       Create it and place all student .docx files inside.")
        sys.exit(1)

    docx_files = sorted(glob.glob(str(folder / "*.docx")))
    if not docx_files:
        print(f"ERROR: No .docx files found in '{STUDENTS_FOLDER}/'")
        sys.exit(1)

    print(f"\n📂  Found {len(docx_files)} file(s) in '{STUDENTS_FOLDER}/'")
    print("-" * 56)

    rows = []
    skipped_files = []
    for fp in docx_files:
        print(f"  Reading: {Path(fp).name} ...", end=" ")
        data = parse_student_doc(fp)
        if data:
            roll = data.get("roll_number", "").strip()
            if not roll:
                reason = "Roll Number is blank"
                print(f"⚠  Skipped — {reason}")
                skipped_files.append((Path(fp).name, reason))
                continue
            rows.append(data)
            name = data.get("student_name") or "(no name)"
            print(f"✓  Roll={roll}  Name={name}")

    if not rows:
        print("\nNo data extracted. Exiting.")
        sys.exit(1)

    output_dir = Path(OUTPUT_FOLDER)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / OUTPUT_EXCEL

    build_excel(rows, output_path)
    print("-" * 56)
    print(f"\n✅  Excel written: {output_path}")
    print(f"   {len(rows)} student row(s) ready for marks entry.")
    if skipped_files:
        print(f"\n⚠  {len(skipped_files)} file(s) skipped:")
        for name, reason in skipped_files:
            print(f"   • {name} — {reason}")
    print("\n📝  Next steps:")
    print("   1. Open marks.xlsx")
    print("   2. Fill in the yellow 'Awarded' columns + feedback columns")
    print("   3. Run:  python fill_marks_and_export.py")


if __name__ == "__main__":
    main()
