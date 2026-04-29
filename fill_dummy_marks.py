"""
================================================================================
  fill_dummy_marks.py
  CA3 Marks Processing — Dummy Marks Filler
================================================================================

PURPOSE
-------
Fills the yellow 'Awarded' and Feedback columns in CA3_Marks/marks.xlsx with
realistic dummy data so you can test fill_marks_and_export.py without manually
entering marks.

  Awarded marks  : random value between 60 % and 100 % of each question's
                   allotted marks (rounded to nearest integer).
  Feedback cells : randomly selected from the pools defined below.

USAGE
-----
    python fill_dummy_marks.py

Run AFTER extract_to_excel.py has produced CA3_Marks/marks.xlsx.
================================================================================
"""

import sys
import random
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not found. Run:  pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

MARKS_EXCEL     = "CA3_Marks/marks.xlsx"
QUESTION_LABELS = ["1a", "1b", "1c", "1d", "1e", "2", "3", "4", "5", "6", "7"]

# Awarded marks are drawn uniformly from [MIN_FRACTION, 1.0] × allotted marks.
MIN_FRACTION = 0.6

# Fixed per-question values (same for all students)
QUESTION_COURSE_OUTCOMES = {
    "1a": "CO1", "1b": "CO1", "1c": "CO2", "1d": "CO2", "1e": "CO2",
    "2":  "CO3", "3":  "CO3", "4":  "CO4", "5":  "CO4", "6":  "CO5", "7": "CO5",
}

QUESTION_BLOOMS_LEVELS = {
    "1a": "L1", "1b": "L1", "1c": "L2", "1d": "L2", "1e": "L3",
    "2":  "L3", "3":  "L4", "4":  "L4", "5":  "L5", "6":  "L5", "7": "L6",
}

QUESTION_AR_REFERENCES = {
    "1a": "Unit 1, Ch.1", "1b": "Unit 1, Ch.1", "1c": "Unit 1, Ch.2",
    "1d": "Unit 1, Ch.2", "1e": "Unit 2, Ch.1",
    "2":  "Unit 2, Ch.2", "3":  "Unit 3, Ch.1", "4":  "Unit 3, Ch.2",
    "5":  "Unit 4, Ch.1", "6":  "Unit 4, Ch.2", "7":  "Unit 5, Ch.1",
}

# Per-student per-question (random)
REMARKS_POOL = [
    "Correct and well explained.",
    "Partially correct; key steps missing.",
    "Incorrect approach; refer to notes.",
    "Satisfactory; minor errors.",
    "Excellent; all edge cases addressed.",
]

STRENGTHS_POOL = [
    "Good understanding of core concepts with clear and structured answers.",
    "Excellent problem-solving approach; solutions are well-reasoned and concise.",
    "Strong command of theoretical concepts with accurate examples.",
    "Demonstrates solid grasp of the subject; answers are precise and well-organised.",
    "Very good analytical thinking with all steps clearly explained.",
]

AREAS_POOL = [
    "Needs to improve time management; some answers were incomplete.",
    "Should focus on revising edge cases and boundary conditions.",
    "More practice on complex data structure problems is recommended.",
    "Deeper understanding of algorithmic complexity (Big-O) is needed.",
    "Handwriting and presentation could be improved for clarity.",
]

CORRECTIVE_POOL = [
    "Revise chapters 4 and 5 and attempt additional practice problems.",
    "Refer to recommended textbook exercises on sorting and searching.",
    "Attend extra tutorial sessions for algorithm analysis.",
    "Practice writing pseudocode and trace tables for complex problems.",
    "Review lecture notes on trees and graphs; attempt past papers.",
]

# ── styles ────────────────────────────────────────────────────────────────────
_YELLOW = PatternFill("solid", fgColor="FFFACD")


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def header_map(ws):
    """Return {header_string: column_index (1-based)} from row 1."""
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def allotted_hdr(label):        return f"Allotted {label.upper()}"
def awarded_hdr(label):         return f"Awarded {label.upper()}"
def course_outcome_hdr(label):  return f"Course Outcome {label.upper()}"
def blooms_level_hdr(label):    return f"Blooms Level {label.upper()}"
def remarks_hdr(label):         return f"Remarks {label.upper()}"
def ar_reference_hdr(label):    return f"AR Ref {label.upper()}"


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    excel_path = Path(MARKS_EXCEL)
    if not excel_path.exists():
        print(f"ERROR: '{MARKS_EXCEL}' not found.")
        print("       Run  python extract_to_excel.py  first.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    col = header_map(ws)

    # Verify required columns are present
    required = (
        [allotted_hdr(q)       for q in QUESTION_LABELS]
        + [awarded_hdr(q)      for q in QUESTION_LABELS]
        + [course_outcome_hdr(q) for q in QUESTION_LABELS]
        + [blooms_level_hdr(q) for q in QUESTION_LABELS]
        + [remarks_hdr(q)      for q in QUESTION_LABELS]
        + [ar_reference_hdr(q) for q in QUESTION_LABELS]
        + ["Strengths", "Areas for Improvement", "Corrective Measures"]
    )
    missing = [h for h in required if h not in col]
    if missing:
        print(f"ERROR: Expected columns not found: {missing}")
        sys.exit(1)

    data_rows = ws.max_row - 1
    if data_rows < 1:
        print("No student rows found in the spreadsheet.")
        sys.exit(0)

    print(f"\nFilling dummy marks for {data_rows} student(s) in '{MARKS_EXCEL}' ...")
    print("-" * 60)

    for row_num in range(2, ws.max_row + 1):
        # ── awarded marks ─────────────────────────────────────────────────
        for q in QUESTION_LABELS:
            allotted_val = ws.cell(row=row_num, column=col[allotted_hdr(q)]).value
            try:
                allotted = int(str(allotted_val).strip())
            except (ValueError, TypeError):
                allotted = 0

            if allotted > 0:
                raw = random.uniform(MIN_FRACTION, 1.0) * allotted
                awarded = int(raw) if raw == int(raw) else round(raw, 1)
            else:
                awarded = 0

            cell = ws.cell(row=row_num, column=col[awarded_hdr(q)], value=awarded)
            cell.fill      = _YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── course outcome & bloom's level (fixed per question) ───────────
        _TEAL = PatternFill("solid", fgColor="E0F4F4")
        for q in QUESTION_LABELS:
            c = ws.cell(row=row_num, column=col[course_outcome_hdr(q)],
                        value=QUESTION_COURSE_OUTCOMES[q])
            c.fill = _TEAL
            c.alignment = Alignment(horizontal="center", vertical="center")

            c = ws.cell(row=row_num, column=col[blooms_level_hdr(q)],
                        value=QUESTION_BLOOMS_LEVELS[q])
            c.fill = _TEAL
            c.alignment = Alignment(horizontal="center", vertical="center")

        # ── remarks (random per student per question) ─────────────────────
        for q in QUESTION_LABELS:
            c = ws.cell(row=row_num, column=col[remarks_hdr(q)],
                        value=random.choice(REMARKS_POOL))
            c.fill      = _YELLOW
            c.alignment = Alignment(vertical="center", wrap_text=True)

        # ── AR reference (fixed per question) ─────────────────────────────
        for q in QUESTION_LABELS:
            c = ws.cell(row=row_num, column=col[ar_reference_hdr(q)],
                        value=QUESTION_AR_REFERENCES[q])
            c.fill      = _YELLOW
            c.alignment = Alignment(horizontal="center", vertical="center")

        # ── feedback ──────────────────────────────────────────────────────
        feedback = [
            ("Strengths",              STRENGTHS_POOL),
            ("Areas for Improvement",  AREAS_POOL),
            ("Corrective Measures",    CORRECTIVE_POOL),
        ]
        for hdr, pool in feedback:
            cell = ws.cell(row=row_num, column=col[hdr], value=random.choice(pool))
            cell.fill      = _YELLOW
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        upid = ws.cell(row=row_num, column=col.get("UPID", 1)).value or ""
        name = ws.cell(row=row_num, column=col.get("Student Name", 2)).value or ""
        print(f"  ✓  {str(upid):<16}  {name}")

    wb.save(excel_path)
    print("-" * 60)
    print(f"\nDone! Saved to: {excel_path}")
    print("\nNext step:  python fill_marks_and_export.py")


if __name__ == "__main__":
    main()
